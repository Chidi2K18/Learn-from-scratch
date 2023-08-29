from openpyxl import Workbook, load_workbook
import json

wb = Workbook()
workbook = load_workbook(filename="Python/sampledataworkorders.xlsx")

#workbook.sheetnames

sheet = workbook.active

sheet["A1"] = "HEllo"
sheet["B1"] = "World"
sheet["C1"] = "This isnt as easy as I remember"

wb.save(filename="hallo world.xlsx")

test_sheet = wb.create_sheet("test")

test_sheet.title = "Tester Title"
#print a range of cells by column
print(sheet['A1:C1'])
#print a range of cells by row
print(sheet[1:2])
#print a column of cells
print(sheet['A'])
#print a row of cells
print(sheet[1])
#print a specific cell using its column and row as reference
print(sheet.cell(row=1, column=2).value)
print(sheet.cell(row=1, column=3).value)

products = {}
#loop through the sheet and return between the first 2 rows and 3 columns
#you can drop the parameters if you want to go through the whole dataset
# you can create objects from the data in your spreadsheets
# for row in sheet.iter_rows(min_row=2,
#                            min_col=4,
#                            max_col=7,
#                            values_only=True):
#     product_id = row[0]
#     product = {
#         "parent": row[1],
#         "title": row[2],
#         "category": row[3]
#     }
#     products[product_id] = product

# print(json.dumps(products))